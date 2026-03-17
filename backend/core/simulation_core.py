import pandas as pd
import os
import time
import functools
import duckdb
import json
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import datetime, timedelta

# Configuración de rutas para RPK NEXUS
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, "db", "rpk_analytical.duckdb")

try:
    from backend.db import models_sim as database
except ImportError:
    import sys
    sys.path.append(os.path.dirname(BASE_DIR))
    from backend.db import models_sim as database

# Variable global para cachear el DataFrame
_df_cache = None

def time_it(func):
    """Decorador para medir el tiempo de ejecución de las funciones."""
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        start_time = time.perf_counter()
        result = func(*args, **kwargs)
        end_time = time.perf_counter()
        print(f"[PERF] {func.__name__} tardó {end_time - start_time:.4f} segundos")
        return result
    return wrapper

def get_base_dataframe():
    """Retorna una copia del DataFrame maestro leyendo desde DuckDB (Carril B)."""
    global _df_cache
    
    # Auto-invalidar cache si faltan columnas críticas (tras actualizaciones en caliente)
    if _df_cache is not None:
        required_cols = {'UATC', 'Fase', 'Articulo', 'Centro'}
        if not required_cols.issubset(set(_df_cache.columns)):
            print("[WARN] Cache inválido: faltan columnas críticas. Recargando desde DuckDB...", flush=True)
            _df_cache = None
        else:
            return _df_cache.copy()

    try:
        print(f"[INFO] Cargando Maestro desde DuckDB (Sincronización v5.9.3)...", flush=True)
        start_load = time.perf_counter()
        
        with duckdb.connect(DB_PATH) as conn:
            # Seleccionamos las columnas necesarias mapeándolas a los nombres esperados por el simulador.
            # 1. Piezas_Hora (del Excel) se mapea a "Piezas por minuto" dividiendo por 60.
            # 2. Si OEE es 0 o NaN en el Excel, aplicamos un fallback de 0.75 (75%) para evitar explosión de saturación.
            query = """
                SELECT 
                    Articulo, 
                    Centro, 
                    CAST(Piezas_Hora / 60.0 AS DOUBLE) as "Piezas por minuto",
                    CASE 
                        WHEN OEE IS NULL OR OEE <= 0 THEN 0.75 
                        ELSE CAST(OEE AS DOUBLE) 
                    END as "%OEE",
                    CAST(Volumen_Anual AS DOUBLE) as "Volumen anual",
                    Fase,
                    UATC,
                    Centro as centro_original
                FROM maestro_fleje
                WHERE TRY_CAST(Centro AS INTEGER) BETWEEN 100 AND 999
            """
            _df_cache = conn.execute(query).df()
            
            # Limpieza y tipado
            _df_cache['Articulo'] = _df_cache['Articulo'].astype(str).str.strip()
            _df_cache['Centro'] = _df_cache['Centro'].astype(str).str.strip()
            
            # Columnas de arquitectura Nexus que no están en el Parquet analítico
            _df_cache['dias laborales 2026'] = 238
            _df_cache['Setup (h)'] = 0.0
            _df_cache['Ratio_MOD'] = 1.0

        end_load = time.perf_counter()
        print(f"[OK] {len(_df_cache)} rutas maestras cargadas desde DuckDB en {end_load - start_load:.4f}s.", flush=True)
        
    except Exception as e:
        print(f"[ERROR] Error al cargar DataFrame desde DuckDB: {e}")
        return None
        
    return _df_cache.copy()

def _log_pending_excel_update(msg: str):
    try:
        pending_file = os.path.join(BASE_DIR, "data_lake", "pending_excel_updates.json")
        if not os.path.exists(os.path.dirname(pending_file)):
            os.makedirs(os.path.dirname(pending_file), exist_ok=True)
        import json
        event = {"timestamp": datetime.now().isoformat(), "msg": msg}
        with open(pending_file, "a") as f:
            f.write(json.dumps(event) + "\n")
    except:
        pass

def add_article(data: dict) -> dict:
    import openpyxl
    df = get_base_dataframe()
    if df is None:
        return {"status": "error", "message": "No se pudo cargar el origen de datos."}

    articulo = str(data.get("articulo")).strip()
    centro = str(data.get("centro")).strip()
    
    mask = (df['Articulo'].astype(str) == articulo) & (df['Centro'].astype(str) == centro)
    if not df[mask].empty:
        return {"status": "error", "message": f"El artículo {articulo} ya existe en el centro {centro}."}

    # Parámetros por si queremos reescribir la fila entera correctamente en Excel
    dias_lab = float(data.get("dias_laborales", 238))
    vol_anual = float(data.get("volumen_anual", 0))
    ppm = float(data.get("piezas_por_minuto", 0))
    oee = float(data.get("oee", 0.75))

    # Nueva fila de Pandas para la Caché en vivo
    nueva_fila = {
        'Articulo': articulo,
        'Centro': centro,
        'Volumen anual': vol_anual,
        'Piezas por minuto': ppm,
        '%OEE': oee,
        'dias laborales 2026': dias_lab,
        'Setup (h)': 0.0,
        'Ratio_MOD': 1.0,
        'centro_original': centro
    }
    
    # 1. Update in-memory DB (Pickle cache deprecated in v5.9.3 for simulator, we use DuckDB)
    global _df_cache
    _df_cache = pd.concat([_df_cache, pd.DataFrame([nueva_fila])], ignore_index=True)

    # 2. Try updating the physical Excel (Mantener como backup de histórico manual)
    try:
        EXCEL_PATH = os.path.join(BASE_DIR, "db", "MAESTRO FLEJE.xlsx")
        pph = ppm * 60
        ppd_16 = pph * 16
        ppd_24 = pph * 24
        ppd_oee_24 = ppd_24 * oee
        ppd_oee_16 = ppd_16 * oee
        pps_24 = ppd_24 * 5
        pps_16 = ppd_16 * 5

        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        new_row = [dias_lab, articulo, float(centro), vol_anual, ppm, pph, ppd_16, ppd_24, ppd_oee_24, ppd_oee_16, pps_24, pps_16, oee]
        ws.append(new_row)
        wb.save(EXCEL_PATH)
        wb.close()
        return {"status": "success", "message": "Memoria y Excel actualizados correctamente."}
        
    except PermissionError:
        print("[WARN] Excel bloqueado al añadir. Escribiendo en pending.")
        _log_pending_excel_update(f"ADD: {articulo} en {centro}")
        return {"status": "warning", "message": "Memoria actualizada, pero Excel maestro está bloqueado. Se sincronizará más tarde."}
    except Exception as e:
        print(f"[ERROR] add_article excel mutator: {e}")
        return {"status": "warning", "message": f"Memoria actualizada. Error archivo físico: {str(e)}"}

def delete_article(articulo: str, centro: str) -> dict:
    import openpyxl
    df = get_base_dataframe()
    if df is None:
        return {"status": "error", "message": "No se pudo cargar el origen de datos."}
        
    articulo = str(articulo).strip()
    centro = str(centro).strip()
    
    mask = (df['Articulo'].astype(str) == articulo) & (df['Centro'].astype(str) == centro)
    if not mask.any():
        return {"status": "error", "message": f"Artículo no encontrado ({articulo} en {centro})."}
        
    # 1. Update in-memory
    global _df_cache
    _df_cache = df[~mask].copy()

    # 2. Try updating the physical Excel
    try:
        EXCEL_PATH = os.path.join(BASE_DIR, "db", "MAESTRO FLEJE.xlsx")
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        
        row_to_delete = None
        for row in range(2, ws.max_row + 1):
            art_val = str(ws.cell(row=row, column=2).value).strip().replace('.0', '')
            cen_val = str(ws.cell(row=row, column=3).value).strip().replace('.0', '')
            
            if art_val == articulo and cen_val == centro:
                row_to_delete = row
                break
                
        if row_to_delete is not None:
            ws.delete_rows(row_to_delete, 1)
            wb.save(EXCEL_PATH)
            
        wb.close()
        return {"status": "success", "message": "Memoria y Excel actualizados correctamente."}
        
    except PermissionError:
        _log_pending_excel_update(f"DELETE: {articulo} en {centro}")
        return {"status": "warning", "message": "Memoria actualizada, pero Excel maestro está bloqueado. Se sincronizará más tarde."}
    except Exception as e:
        return {"status": "warning", "message": f"Memoria actualizada. Error archivo físico: {str(e)}"}


@time_it
def calculate_saturation(df: pd.DataFrame, dias_laborales_override: int = None, horas_turno_default: int = 16):
    """
    Calcula la saturación basada en las columnas proyectadas desde DuckDB.
    """
    # Aseguramos tipos de datos
    df['Volumen anual'] = pd.to_numeric(df['Volumen anual'], errors='coerce').fillna(0)
    df['Piezas por minuto'] = pd.to_numeric(df['Piezas por minuto'], errors='coerce').fillna(0)
    df['%OEE'] = pd.to_numeric(df['%OEE'], errors='coerce').fillna(0.75)
    
    if 'horas_turno' not in df.columns:
        df['horas_turno'] = horas_turno_default
    
    if dias_laborales_override is not None:
        df['dias laborales 2026'] = dias_laborales_override
    else:
        df['dias laborales 2026'] = pd.to_numeric(df['dias laborales 2026'], errors='coerce').fillna(238)

    if 'Setup (h)' not in df.columns:
        df['Setup (h)'] = 0.0
    else:
        df['Setup (h)'] = pd.to_numeric(df['Setup (h)'], errors='coerce').fillna(0)

    if 'Ratio_MOD' not in df.columns:
        df['Ratio_MOD'] = 1.0
    else:
        df['Ratio_MOD'] = pd.to_numeric(df['Ratio_MOD'], errors='coerce').fillna(1.0)

    # Cálculos dinámicos
    df['Piezas por hora'] = df['Piezas por minuto'] * 60
    
    # Normalización OEE (0-1)
    df_oee_calc = df['%OEE'].copy()
    df_oee_calc = df_oee_calc.apply(lambda x: x/100.0 if x > 1.1 else x)
    # Evitar divisiones por cero críticas
    df_oee_calc = df_oee_calc.clip(lower=0.01)

    denominador = (df['Piezas por hora'] * df_oee_calc)
    df['Horas_Produccion'] = (df['Volumen anual'] / denominador).replace([float('inf'), -float('inf')], 0).fillna(0)
    df['Horas_Totales'] = df['Horas_Produccion'] + df['Setup (h)']
    df['Horas_Hombre'] = (df['Horas_Produccion'] * df['Ratio_MOD'].fillna(1.0)) + df['Setup (h)']
    
    df['Capacidad_Anual_H'] = df['dias laborales 2026'] * df['horas_turno']
    df['Saturacion'] = (df['Horas_Totales'] / df['Capacidad_Anual_H']).replace([float('inf'), -float('inf')], 0).fillna(0)

    total_horas_global = df['Horas_Totales'].sum()
    df['Impacto'] = (df['Horas_Totales'] / total_horas_global) if total_horas_global > 0 else 0.0
    
    return df

def get_pedidos_actuales(dias_laborales: int = None):
    try:
        pedidos_dir = os.path.join(BASE_DIR, "data_lake", "transaccional", "pedidos")
        if not os.path.exists(pedidos_dir):
            return None

        years = sorted([d for d in os.listdir(pedidos_dir) if d.startswith("year=")], reverse=True)
        if not years: return None
        months = sorted([d for d in os.listdir(os.path.join(pedidos_dir, years[0])) if d.startswith("month=")], reverse=True)
        if not months: return None
        
        day_dir = os.path.join(pedidos_dir, years[0], months[0])
        files = sorted([f for f in os.listdir(day_dir) if f.endswith(".parquet")], reverse=True)
        if not files: return None
        
        latest_parquet = os.path.join(day_dir, files[0])
        df_orders = pd.read_parquet(latest_parquet)
        df_orders.columns = [str(c).strip().upper() for c in df_orders.columns]

        if dias_laborales is not None:
            horizonte = datetime.now() + timedelta(days=int(dias_laborales))
            if 'FECHA_ENTREGA' in df_orders.columns:
                df_orders['FECHA_ENTREGA'] = pd.to_datetime(df_orders['FECHA_ENTREGA'], errors='coerce')
                df_orders = df_orders[(df_orders['FECHA_ENTREGA'] <= horizonte) | (df_orders['FECHA_ENTREGA'].isna())].copy()

        if 'CANT_PENDIENTE' in df_orders.columns:
            df_orders['CANT_PENDIENTE'] = pd.to_numeric(df_orders['CANT_PENDIENTE'], errors='coerce').fillna(0)
            df_demanda = df_orders.groupby('ARTICULO')['CANT_PENDIENTE'].sum().reset_index()
            df_demanda.columns = ['ARTICULO_lake', 'DEMANDA_ACTUAL']
            return df_demanda
        
        return None
    except Exception as e:
        print(f"[ERROR] Error en get_pedidos_actuales: {e}")
        return None

def get_stock_actual():
    try:
        stock_dir = os.path.join(BASE_DIR, "data_lake", "transaccional", "existencias")
        if not os.path.exists(stock_dir):
            return None
        
        years = sorted([d for d in os.listdir(stock_dir) if d.startswith("year=")], reverse=True)
        if not years: return None
        months = sorted([d for d in os.listdir(os.path.join(stock_dir, years[0])) if d.startswith("month=")], reverse=True)
        if not months: return None
        
        day_dir = os.path.join(stock_dir, years[0], months[0])
        files = sorted([f for f in os.listdir(day_dir) if f.endswith(".parquet")], reverse=True)
        if not files: return None
        
        latest_parquet = os.path.join(day_dir, files[0])
        df_stock = pd.read_parquet(latest_parquet)
        df_stock.columns = [str(c).strip().upper() for c in df_stock.columns]
        
        if 'ARTICULO' in df_stock.columns and 'CANTIDAD' in df_stock.columns:
            df_stock['CANTIDAD'] = pd.to_numeric(df_stock['CANTIDAD'], errors='coerce').fillna(0)
            df_res = df_stock.groupby('ARTICULO')['CANTIDAD'].sum().reset_index()
            df_res.columns = ['ARTICULO_lake', 'STOCK_ACTUAL']
            return df_res
        
        return None
    except Exception as e:
        print(f"[ERROR] Error en get_stock_actual: {e}")
        return None

@time_it
def get_simulation_data(db: Session, scenario_id: int = None, dias_laborales: int = None, overrides_list: List = None, horas_turno: int = None, center_configs: dict = None, use_actual: bool = False):
    df = get_base_dataframe()
    if df is None:
        return {"detail": [], "summary": [], "meta": {}}
    
    if use_actual:
        df_pedidos = get_pedidos_actuales(dias_laborales)
        df_stock = get_stock_actual()
        
        if df_pedidos is not None:
            df = df.merge(df_pedidos, left_on='Articulo', right_on='ARTICULO_lake', how='left')
            df['DEMANDA_ACTUAL'] = df['DEMANDA_ACTUAL'].fillna(0)
            if df_stock is not None:
                df = df.merge(df_stock, left_on='Articulo', right_on='ARTICULO_lake', how='left')
                df['STOCK_ACTUAL'] = df['STOCK_ACTUAL'].fillna(0)
                df['Volumen anual'] = (df['DEMANDA_ACTUAL'] - df['STOCK_ACTUAL']).clip(lower=0)
                df = df.drop(columns=['ARTICULO_lake_x', 'ARTICULO_lake_y', 'DEMANDA_ACTUAL', 'STOCK_ACTUAL'], errors='ignore')
            else:
                df['Volumen anual'] = df['DEMANDA_ACTUAL']
                df = df.drop(columns=['ARTICULO_lake', 'DEMANDA_ACTUAL'], errors='ignore')
        else:
            df['Volumen anual'] = 0
    
    h_turno = int(horas_turno) if horas_turno is not None else 16
    df['horas_turno'] = h_turno
    
    if center_configs:
        for centro, config in center_configs.items():
            mask_c = df['Centro'].astype(str) == str(centro)
            if isinstance(config, dict):
                if 'shifts' in config:
                    df.loc[mask_c, 'horas_turno'] = int(config['shifts'])
                if 'personnel_ratio' in config:
                    df.loc[mask_c, 'Ratio_MOD'] = float(config['personnel_ratio'])
    
    selected_overrides = []
    if scenario_id:
        selected_overrides = list(db.query(database.ScenarioDetail).filter(database.ScenarioDetail.scenario_id == scenario_id).all())
    if overrides_list:
        selected_overrides = selected_overrides + list(overrides_list)

    for ov in selected_overrides:
        art = getattr(ov, 'articulo', None) or (ov.articulo if hasattr(ov, 'articulo') else None)
        cen = getattr(ov, 'centro', None) or (ov.centro if hasattr(ov, 'centro') else None)
        mask = (df['Articulo'].astype(str) == str(art)) & (df['Centro'].astype(str) == str(cen))
        
        oee = getattr(ov, 'oee_override', None)
        ppm = getattr(ov, 'ppm_override', None)
        dem = getattr(ov, 'demanda_override', None)
        nc = getattr(ov, 'new_centro', None)
        ht = getattr(ov, 'horas_turno_override', None)

        if oee is not None: df.loc[mask, '%OEE'] = oee
        if ppm is not None: df.loc[mask, 'Piezas por minuto'] = ppm
        if dem is not None: df.loc[mask, 'Volumen anual'] = dem
        if nc is not None: df.loc[mask, 'Centro'] = nc
        if ht is not None: df.loc[mask, 'horas_turno'] = ht
        if hasattr(ov, 'personnel_ratio_override') and ov.personnel_ratio_override is not None:
            df.loc[mask, 'Ratio_MOD'] = ov.personnel_ratio_override
        if getattr(ov, 'setup_time_override', None) is not None: 
            df.loc[mask, 'Setup (h)'] = ov.setup_time_override

    d_lab = int(dias_laborales) if dias_laborales is not None else None
    df = calculate_saturation(df, d_lab, h_turno)
    
    centro_summary = df.groupby('Centro').agg({
        'Saturacion': 'sum',
        'Volumen anual': 'sum',
        'Horas_Totales': 'sum',
        'Horas_Hombre': 'sum',
        'Articulo': 'count'
    }).reset_index()
    centro_summary.rename(columns={'Articulo': 'Num_Articulos'}, inplace=True)
    
    # fillna SOLO en columnas numéricas para no corromper UATC (string) ni Fase
    numeric_cols = df.select_dtypes(include='number').columns
    df[numeric_cols] = df[numeric_cols].fillna(0).replace([float('inf'), -float('inf')], 0)
    
    # Rellenar NaNs en columnas de texto con string vacío para evitar fallos de serialización JSON en FastAPI
    non_numeric_cols = df.select_dtypes(exclude='number').columns
    df[non_numeric_cols] = df[non_numeric_cols].fillna("")
    
    numeric_cols_sum = centro_summary.select_dtypes(include='number').columns
    centro_summary[numeric_cols_sum] = centro_summary[numeric_cols_sum].fillna(0).replace([float('inf'), -float('inf')], 0)
    
    non_numeric_cols_sum = centro_summary.select_dtypes(exclude='number').columns
    centro_summary[non_numeric_cols_sum] = centro_summary[non_numeric_cols_sum].fillna("")

    # Aseguramos que UATC y Fase se incluyan en el detalle para el frontend
    cols_to_return = [
        'Articulo', 'Centro', 'Volumen anual', 'Piezas por minuto', 
        '%OEE', 'Saturacion', 'Ratio_MOD', 'UATC', 'Fase',
        'Horas_Produccion', 'Horas_Totales', 'Horas_Hombre', 'Setup (h)', 'dias laborales 2026'
    ]
    # Filtrar solo columnas existentes para evitar errores si alguna falta por alguna razón
    cols_final = [c for c in cols_to_return if c in df.columns]

    return {
        "detail": df[cols_final].to_dict(orient="records"),
        "summary": centro_summary.to_dict(orient="records"),
        "meta": {
            "dias_laborales": d_lab if d_lab is not None else 238,
            "horas_turno_global": h_turno,
            "center_configs": center_configs or {},
            "applied_overrides": [
                {
                    "articulo": getattr(ov, 'articulo', None) or (ov.articulo if hasattr(ov, 'articulo') else None),
                    "centro": getattr(ov, 'centro', None) or (ov.centro if hasattr(ov, 'centro') else None),
                    "oee_override": getattr(ov, 'oee_override', None),
                    "ppm_override": getattr(ov, 'ppm_override', None),
                    "demanda_override": getattr(ov, 'demanda_override', None),
                    "new_centro": getattr(ov, 'new_centro', None),
                    "horas_turno_override": getattr(ov, 'horas_turno_override', None),
                    "personnel_ratio_override": getattr(ov, 'personnel_ratio_override', None),
                    "setup_time_override": getattr(ov, 'setup_time_override', None)
                } for ov in selected_overrides
            ] if selected_overrides else []
        }
    }
